import os
if not hasattr(os, "add_dll_directory"):
    os.add_dll_directory = lambda x: None

import logging
import azure.functions as func

import sys
import json
import base64
import cv2
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

# Blob Storage 관련 import 추가
from azure.storage.blob import BlobServiceClient, ContentSettings

# ---------------------------
# 전역: 색상 범위 (HSV)
# ---------------------------
color_ranges = {
    "검정":  ((0,   0,   0),   (10, 255,  50)),
    "갈색":  ((10,  80,  5),   (30, 255, 255)),
    "초록":  ((60,  50,  50),  (85, 255, 255)),
    "보라":  ((130, 50,  50),  (170, 255, 255)),
    "파랑":  ((90,  50,  50),  (114, 255, 255)),

    "나무둘레": ((115, 50, 50),  (129, 255, 255)),
    "나무표시": ((32,  50, 50),  (59, 255, 255)),
    "나무표시2":((3,  240,150),  (10, 255, 255))
}

# 등급 카운트 (전역)
grade_count = {"A":0, "B":0, "C":0, "D":0, "E":0}

def calc_grade(ratio: float) -> str:
    """
    초록+보라+파랑 비율(ratio)에 따라 등급 산정
    """
    if 0 <= ratio < 1:
        return "A"
    elif 1 <= ratio <= 19:
        return "B"
    elif 20 <= ratio <= 39:
        return "C"
    elif 40 <= ratio <= 49:
        return "D"
    else:
        return "E"

def analyze_one_image(tree_id: str, image_path: str):
    """
    단일 이미지를 분석하여 색상 픽셀수/등급 정보를 반환.
    실패 시 None.
    """
    if not os.path.exists(image_path):
        logging.info(f"[오류] 파일 없음: {image_path}")
        return None

    img_bgr = cv2.imread(image_path)
    if img_bgr is None:
        logging.info(f"[오류] OpenCV로 읽지 못함: {image_path}")
        return None

    # BGR → HSV 변환
    img_hsv = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2HSV)

    # 나무둘레 + 표시 + 표시2 합치기
    outer_mask = np.zeros(img_hsv.shape[:2], dtype=np.uint8)
    for key in ["나무둘레", "나무표시", "나무표시2"]:
        lo, up = color_ranges[key]
        tmp = cv2.inRange(img_hsv, lo, up)
        outer_mask = cv2.bitwise_or(outer_mask, tmp)

    # 모폴로지
    kernel = np.ones((3,3), np.uint8)
    outer_mask = cv2.dilate(outer_mask, kernel, iterations=1)
    outer_mask = cv2.erode(outer_mask, kernel, iterations=1)

    # 컨투어 찾기
    contours, _ = cv2.findContours(outer_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        logging.info(f"[결과] 외곽 컨투어 없음: {image_path}")
        return None

    largest = max(contours, key=cv2.contourArea)
    roi_mask = np.zeros(img_bgr.shape[:2], dtype=np.uint8)
    cv2.drawContours(roi_mask, [largest], -1, 255, -1)

    # 5색
    target_colors = ["검정", "갈색", "초록", "보라", "파랑"]
    color_counts = {}
    for c in target_colors:
        lo, up = color_ranges[c]
        mask_c = cv2.inRange(img_hsv, lo, up)
        final_mask = cv2.bitwise_and(mask_c, mask_c, mask=roi_mask)
        cnt = cv2.countNonZero(final_mask)
        color_counts[c] = cnt

    sum_of_5 = sum(color_counts.values())
    if sum_of_5 == 0:
        logging.info(f"[결과] 5색 픽셀 없음: {image_path}")
        return None

    # 검정+갈색
    black_brown = color_counts["검정"] + color_counts["갈색"]
    black_brown_ratio = round((black_brown / sum_of_5)*100, 2)

    # 초록+보라+파랑
    gpb = color_counts["초록"] + color_counts["보라"] + color_counts["파랑"]
    gpb_ratio = round((gpb / sum_of_5)*100, 2)

    overall_grade = calc_grade(gpb_ratio)

    return {
        "tree_id": tree_id,
        "image_path": image_path,
        "color_counts": color_counts,
        "sum_of_5": sum_of_5,
        "black_brown_count": black_brown,
        "black_brown_ratio": black_brown_ratio,
        "green_purple_blue_count": gpb,
        "green_purple_blue_ratio": gpb_ratio,
        "overall_grade": overall_grade
    }

def analyze_multiple_images(image_list, excel_filename="analysis.xlsx"):
    """
    여러 이미지를 한 번에 분석 → 하나의 엑셀로 작성.
    image_list: [(수목번호, 이미지경로), (...), ...]
    """
    global grade_count
    grade_count = {"A":0, "B":0, "C":0, "D":0, "E":0}

    # 기존 파일 있으면 삭제
    if os.path.exists(excel_filename):
        os.remove(excel_filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "분석결과"

    # 1행 병합 → "음파단층촬영 조사현황"
    ws.merge_cells("A1:G1")
    ws["A1"] = "음파단층촬영 조사현황"
    # 2행 헤더
    ws.append(["수목번호", "이미지", "구분", "픽셀수", "합계", "비율(%)", "등급"])

    current_row = 3

    for (tree_id, img_path) in image_list:
        result = analyze_one_image(tree_id, img_path)
        if result is None:
            continue

        color_counts = result["color_counts"]
        sum_of_5     = result["sum_of_5"]
        black_brown_count = result["black_brown_count"]
        black_brown_ratio = result["black_brown_ratio"]
        gpb_count = result["green_purple_blue_count"]
        gpb_ratio = result["green_purple_blue_ratio"]
        overall_grade = result["overall_grade"]

        grade_count[overall_grade] += 1

        start_row = current_row
        end_row   = start_row + 4

        # 수목번호(A열), 이미지경로(B열), 등급(G열)을 병합
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)  # A
        ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)  # B
        ws.merge_cells(start_row=start_row, start_column=7, end_row=end_row, end_column=7)  # G

        ws.cell(row=start_row, column=1, value=tree_id)
        ws.cell(row=start_row, column=2, value=img_path)
        ws.cell(row=start_row, column=7, value=overall_grade)

        # 검정+갈색
        ws.merge_cells(start_row=start_row, start_column=5, end_row=start_row+1, end_column=5)
        ws.merge_cells(start_row=start_row, start_column=6, end_row=start_row+1, end_column=6)
        ws.cell(row=start_row,   column=5, value=black_brown_count)
        ws.cell(row=start_row,   column=6, value=black_brown_ratio)

        # 초록+보라+파랑
        ws.merge_cells(start_row=start_row+2, start_column=5, end_row=end_row, end_column=5)
        ws.merge_cells(start_row=start_row+2, start_column=6, end_row=end_row, end_column=6)
        ws.cell(row=start_row+2, column=5, value=gpb_count)
        ws.cell(row=start_row+2, column=6, value=gpb_ratio)

        # C/D : 5색
        c_list = ["검정", "갈색", "초록", "보라", "파랑"]
        for i, cname in enumerate(c_list):
            r = start_row + i
            ws.cell(row=r, column=3, value=cname)
            ws.cell(row=r, column=4, value=color_counts[cname])

        # (추가) 이미지 삽입
        if os.path.exists(img_path):
            try:
                excel_img = XLImage(img_path)
                excel_img.width = 140
                excel_img.height = 140
                anchor_cell = f"B{start_row}"
                ws.add_image(excel_img, anchor_cell)

                # 행 높이 조절
                row_height_pts = 21
                for rr in range(start_row, end_row+1):
                    ws.row_dimensions[rr].height = row_height_pts

                ws.column_dimensions["B"].width = 17.5

            except Exception as e:
                logging.info(f"[이미지삽입오류] {img_path}: {e}")

        current_row += 5

    # 등급표 (J1~L1 헤더, J2~L6 데이터)
    ws.cell(row=1, column=10, value="등급")
    ws.cell(row=1, column=11, value="기준")
    ws.cell(row=1, column=12, value="수량")

    grade_table = [
        ("A",   "0",      grade_count["A"]),
        ("B",   "1-19",   grade_count["B"]),
        ("C",   "20-39",  grade_count["C"]),
        ("D",   "40-49",  grade_count["D"]),
        ("E",   "50이상", grade_count["E"]),
    ]
    row2 = 2
    for i, (g, criteria, qty) in enumerate(grade_table):
        rr = row2 + i
        ws.cell(row=rr, column=10, value=g)
        ws.cell(row=rr, column=11, value=criteria)
        ws.cell(row=rr, column=12, value=qty)

    # 테두리 & 중앙정렬
    thin_side = Side(style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    last_row_main = current_row - 1
    for row_cells in ws.iter_rows(min_row=1, max_row=last_row_main, min_col=1, max_col=7):
        for cell in row_cells:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    for row_cells in ws.iter_rows(min_row=1, max_row=6, min_col=10, max_col=12):
        for cell in row_cells:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    wb.save(excel_filename)
    logging.info(f"[결과] 전체 {len(image_list)}개 이미지 분석 완료 → {excel_filename}")

def decode_and_run(json_str):
    data = json.loads(json_str)
    image_list = []

    # 최대 15장 예시
    for i in range(1, 16):
        num_key = f"img{i}Num"
        img_key = f"img{i}"
        if num_key in data and img_key in data:
            tree_id = data[num_key]
            b64_str = data[img_key]
            if not b64_str:
                continue
            # 쓰기 권한이 있는 /tmp 디렉터리를 사용합니다.
            local_path = os.path.join("/tmp", f"temp_img{i}.jpg")
            with open(local_path, "wb") as f:
                f.write(base64.b64decode(b64_str))
            image_list.append((tree_id, local_path))

    if not image_list:
        logging.info("[결과] 디코딩된 이미지가 하나도 없습니다.")
        return None

    # JSON에 전달된 "Name" 필드를 추출하여 파일명 생성 (앞뒤 공백 제거)
    name_prefix = data.get("Name", "")
    if name_prefix:
        excel_filename = os.path.join("/tmp", f"{name_prefix}_picus.xlsx")
    else:
        excel_filename = os.path.join("/tmp", "picus.xlsx")
    
    analyze_multiple_images(image_list, excel_filename)
    
    # 생성된 파일명을 반환합니다.
    return excel_filename

# ---------------------------
# 메인 함수 (HTTP Trigger)
# ---------------------------
def main(req: func.HttpRequest) -> func.HttpResponse:
    """
    Azure Function의 HTTP Trigger 엔트리 포인트
    """
    logging.info("Python HTTP trigger function processed a request.")

    # 1) HTTP Body 읽기 (JSON)
    try:
        body_str = req.get_body().decode('utf-8')
    except Exception as e:
        logging.error(f"Error reading request body: {e}")
        return func.HttpResponse("Invalid request body", status_code=400)

    excel_file = decode_and_run(body_str)
    if os.path.exists(excel_file):
        try:
            with open(excel_file, "rb") as f:
                file_data = f.read()

            # Blob Storage 연결 및 업로드
            connect_str = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
            if not connect_str:
                logging.error("AZURE_STORAGE_CONNECTION_STRING not set.")
                return func.HttpResponse("Storage connection string not set", status_code=500)

            blob_service_client = BlobServiceClient.from_connection_string(connect_str)
            container_name = "streetxlsx"  # 실제 컨테이너 이름으로 변경
            blob_path = os.path.basename(excel_file)

            container_client = blob_service_client.get_container_client(container_name)
            try:
                container_client.create_container()
            except Exception as e:
                logging.info(f"Container may already exist: {e}")

            blob_client = container_client.get_blob_client(blob_path)
            content_settings = ContentSettings(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            blob_client.upload_blob(file_data, overwrite=True, content_settings=content_settings)

            return func.HttpResponse(json.dumps({"result": "success"}), status_code=200,headers={"Content-Type": "application/json"})
        except Exception as e:
            logging.error(f"Error uploading file to Blob Storage: {e}")
            return func.HttpResponse("Error uploading file to Blob Storage", status_code=500)
    else:
        return func.HttpResponse("No excel output generated.", status_code=200)

