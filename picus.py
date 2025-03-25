import sys
import json
import base64
import cv2
import numpy as np
import openpyxl
import os
from openpyxl.styles import Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage  # 이미지 삽입용 (Pillow 필요)

# ---------------------------
# 전역: 색상 범위 (HSV)
# ---------------------------
color_ranges = {
    "검정":  ((0,   0,   0),   (10, 255,  50)),
    "갈색":  ((10,  80,  5),   (30, 255, 255)),
    "초록":  ((60,  50,  50),  (85, 255, 255)),
    "보라":  ((130, 50,  50),  (170, 255, 255)),
    "파랑":  ((90,  50,  50),  (114, 255, 255)),

    "나무둘레": ((115, 50, 50),  (129, 255, 255)),
    "나무표시": ((32,  50, 50),  (59, 255, 255)),
    "나무표시2":((3,  240,150),  (10, 255, 255))
}

# 등급 카운트 (전역)
grade_count = {"A":0, "B":0, "C":0, "D":0, "E":0}

def calc_grade(ratio: float) -> str:
    """초록+보라+파랑 비율(ratio)에 따라 등급 산정"""
    if 0 <= ratio < 1:
        return "A"
    elif 1 <= ratio <= 19:
        return "B"
    elif 20 <= ratio <= 39:
        return "C"
    elif 40 <= ratio <= 49:
        return "D"
    else:
        return "E"

def analyze_one_image(tree_id: str, image_path: str):
    """단일 이미지를 분석하여 색상 픽셀수/등급 정보를 반환. 실패 시 None."""
    if not os.path.exists(image_path):
        print(f"[오류] 파일 없음: {image_path}")
        return None

    img_bgr = cv2.imread(image_path)
    if img_bgr is None:
        print(f"[오류] OpenCV로 읽지 못함: {image_path}")
        return None

    img_hsv = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2HSV)

    # 나무둘레+표시+표시2 합치기
    outer_mask = np.zeros(img_hsv.shape[:2], dtype=np.uint8)
    for key in ["나무둘레", "나무표시", "나무표시2"]:
        lo, up = color_ranges[key]
        tmp = cv2.inRange(img_hsv, lo, up)
        outer_mask = cv2.bitwise_or(outer_mask, tmp)

    # 모폴로지
    kernel = np.ones((3,3), np.uint8)
    outer_mask = cv2.dilate(outer_mask, kernel, iterations=1)
    outer_mask = cv2.erode(outer_mask, kernel, iterations=1)

    contours, _ = cv2.findContours(outer_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        print(f"[결과] 외곽 컨투어 없음: {image_path}")
        return None

    largest = max(contours, key=cv2.contourArea)
    roi_mask = np.zeros(img_bgr.shape[:2], dtype=np.uint8)
    cv2.drawContours(roi_mask, [largest], -1, 255, -1)

    # 5색
    target_colors = ["검정", "갈색", "초록", "보라", "파랑"]
    color_counts = {}
    for c in target_colors:
        lo, up = color_ranges[c]
        mask_c = cv2.inRange(img_hsv, lo, up)
        final_mask = cv2.bitwise_and(mask_c, mask_c, mask=roi_mask)
        cnt = cv2.countNonZero(final_mask)
        color_counts[c] = cnt

    sum_of_5 = sum(color_counts.values())
    if sum_of_5 == 0:
        print(f"[결과] 5색 픽셀 없음: {image_path}")
        return None

    # 검정+갈색
    black_brown = color_counts["검정"] + color_counts["갈색"]
    black_brown_ratio = round((black_brown / sum_of_5)*100, 2)

    # 초록+보라+파랑
    gpb = color_counts["초록"] + color_counts["보라"] + color_counts["파랑"]
    gpb_ratio = round((gpb / sum_of_5)*100, 2)

    overall_grade = calc_grade(gpb_ratio)

    return {
        "tree_id": tree_id,
        "image_path": image_path,
        "color_counts": color_counts,
        "sum_of_5": sum_of_5,
        "black_brown_count": black_brown,
        "black_brown_ratio": black_brown_ratio,
        "green_purple_blue_count": gpb,
        "green_purple_blue_ratio": gpb_ratio,
        "overall_grade": overall_grade
    }

def analyze_multiple_images(image_list, excel_filename="analysis_multi.xlsx"):
    """
    여러 이미지를 한 번에 분석 → 하나의 엑셀로 작성.
    image_list: [(수목번호, 이미지경로), (...), ...]
    """
    global grade_count
    grade_count = {"A":0, "B":0, "C":0, "D":0, "E":0}

    # 엑셀 덮어쓰기
    if os.path.exists(excel_filename):
        os.remove(excel_filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "분석결과"

    # 1행 병합 → "음파단층촬영 조사현황"
    ws.merge_cells("A1:G1")
    ws["A1"] = "음파단층촬영 조사현황"

    # 2행 헤더
    ws.append(["수목번호", "이미지", "구분", "픽셀수", "합계", "비율(%)", "등급"])

    current_row = 3

    for (tree_id, img_path) in image_list:
        result = analyze_one_image(tree_id, img_path)
        if result is None:
            continue

        color_counts = result["color_counts"]
        sum_of_5     = result["sum_of_5"]
        black_brown_count = result["black_brown_count"]
        black_brown_ratio = result["black_brown_ratio"]
        gpb_count = result["green_purple_blue_count"]
        gpb_ratio = result["green_purple_blue_ratio"]
        overall_grade = result["overall_grade"]

        grade_count[overall_grade] += 1

        start_row = current_row
        end_row   = start_row + 4

        # A,B,G 병합
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)  # A
        ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)  # B
        ws.merge_cells(start_row=start_row, start_column=7, end_row=end_row, end_column=7)  # G

        ws.cell(row=start_row, column=1, value=tree_id)
        ws.cell(row=start_row, column=2, value=img_path)  # 텍스트로 경로 남김
        ws.cell(row=start_row, column=7, value=overall_grade)

        # E/F 병합 (검정+갈색)
        ws.merge_cells(start_row=start_row,       start_column=5, end_row=start_row+1, end_column=5)
        ws.merge_cells(start_row=start_row,       start_column=6, end_row=start_row+1, end_column=6)
        ws.cell(row=start_row,   column=5, value=black_brown_count)
        ws.cell(row=start_row,   column=6, value=black_brown_ratio)

        # E/F 병합 (초록+보라+파랑)
        ws.merge_cells(start_row=start_row+2,     start_column=5, end_row=end_row, end_column=5)
        ws.merge_cells(start_row=start_row+2,     start_column=6, end_row=end_row, end_column=6)
        ws.cell(row=start_row+2, column=5, value=gpb_count)
        ws.cell(row=start_row+2, column=6, value=gpb_ratio)

        # C/D : 5색
        c_list = ["검정", "갈색", "초록", "보라", "파랑"]
        for i, cname in enumerate(c_list):
            r = start_row + i
            ws.cell(row=r, column=3, value=cname)
            ws.cell(row=r, column=4, value=color_counts[cname])

        # (추가) 이미지 삽입
        if os.path.exists(img_path):
            try:
                excel_img = XLImage(img_path)
                excel_img.width = 140
                excel_img.height = 140
                anchor_cell = f"B{start_row}"
                ws.add_image(excel_img, anchor_cell)

                row_height_pts = 21
                for rr in range(start_row, end_row+1):
                    ws.row_dimensions[rr].height = row_height_pts

                ws.column_dimensions["B"].width = 17.5

            except Exception as e:
                print(f"[이미지삽입오류] {img_path}: {e}")

        current_row += 5

    # 등급표 (J1~L1 헤더, J2~L6 데이터)
    ws.cell(row=1, column=10, value="등급")  
    ws.cell(row=1, column=11, value="기준") 
    ws.cell(row=1, column=12, value="수량") 

    grade_table = [
        ("A",   "0",      grade_count["A"]),
        ("B",   "1-19",   grade_count["B"]),
        ("C",   "20-39",  grade_count["C"]),
        ("D",   "40-49",  grade_count["D"]),
        ("E",   "50이상", grade_count["E"]),
    ]
    row2 = 2
    for i, (g, criteria, qty) in enumerate(grade_table):
        rr = row2 + i
        ws.cell(row=rr, column=10, value=g)
        ws.cell(row=rr, column=11, value=criteria)
        ws.cell(row=rr, column=12, value=qty)

    # 테두리 & 중앙정렬
    thin_side = Side(style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    last_row_main = current_row - 1
    for row_cells in ws.iter_rows(min_row=1, max_row=last_row_main, min_col=1, max_col=7):
        for cell in row_cells:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    for row_cells in ws.iter_rows(min_row=1, max_row=6, min_col=10, max_col=12):
        for cell in row_cells:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    wb.save(excel_filename)
    print(f"[결과] 전체 {len(image_list)}개 이미지 분석 완료 → {excel_filename}")

# ---------------------------------------------------
# 추가 함수: JSON 파라미터를 받아 Base64→파일→리스트
# ---------------------------------------------------
def decode_and_run(json_str):
    data = json.loads(json_str)

    image_list = []
    # img1Num, img1, img2Num, img2, ... 이런 식으로 들어온다고 가정
    for i in range(1, 16):  # 최대 15장
        num_key  = f"img{i}Num"  # 예: "img1Num"
        img_key  = f"img{i}"     # 예: "img1"
        if num_key in data and img_key in data:
            tree_id = data[num_key]
            b64_str = data[img_key]

            if not b64_str:
                # Base64가 비어 있으면 스킵
                continue

            local_path = f"temp_img{i}.jpg"
            with open(local_path, "wb") as f:
                f.write(base64.b64decode(b64_str))

            image_list.append((tree_id, local_path))

    if not image_list:
        print("[결과] 디코딩된 이미지가 하나도 없습니다.")
        return

    excel_out = "analysis.xlsx"
    analyze_multiple_images(image_list, excel_out)


# ---------------------------
# 메인 실행부
# ---------------------------
if __name__ == "__main__":
    # 파라미터가 없으면 종료
    if len(sys.argv) < 2:
        print("[종료] 파라미터가 없어 실행을 종료합니다.")
        import sys
        sys.exit(0)

    # 파라미터가 있으면 decode_and_run
    param_str = sys.argv[1]
    decode_and_run(param_str)